from flask import Blueprint, render_template, jsonify, request, send_file
from utils.database import get_db_connection, DatabaseHandler
from utils.volume_export import export_volume_plan
from utils.volume_ai import generate_volume_suggestions
from openpyxl import Workbook
from io import BytesIO
import datetime

volume_splitter_bp = Blueprint('volume_splitter', __name__)

# --- NEW: mode-specific muscle sets ---
BASIC_MUSCLE_GROUPS = [
    "Neck", "Front-Shoulder", "Rear-Shoulder", "Biceps", "Triceps",
    "Chest", "Forearms", "Abdominals", "Quadriceps", "Hamstrings",
    "Calves", "Latissimus-Dorsi", "Glutes", "Lower Back", "Traps", "Middle-Traps"
]

ADVANCED_MUSCLE_GROUPS = [
    # Deltoids & Trapezius
    "anterior-deltoid", "lateral-deltoid", "posterior-deltoid",
    "upper-trapezius", "traps-middle", "lower-trapezius",
    # Pectorals & Biceps/Triceps
    "upper-pectoralis", "mid-lower-pectoralis",
    "short-head-biceps", "long-head-biceps",
    "lateral-head-triceps", "long-head-triceps", "medial-head-triceps",
    # Forearms
    "wrist-extensors", "wrist-flexors",
    # Core
    "upper-abdominals", "lower-abdominals", "obliques",
    # Hips/Thigh
    "groin", "inner-thigh", "rectus-femoris",
    "inner-quadriceps", "outer-quadriceps",
    # Lower leg
    "soleus", "tibialis", "gastrocnemius",
    # Hamstrings & Glutes
    "medial-hamstrings", "lateral-hamstrings",
    "gluteus-maximus", "gluteus-medius",
    # Back (lats)
    "lats",
    # Spine extensors
    "lowerback"
]


def get_muscle_list_for_mode(mode: str):
    return BASIC_MUSCLE_GROUPS if (mode or "").lower() != "advanced" else ADVANCED_MUSCLE_GROUPS


def build_default_ranges(muscles):
    return {m: {"min": 12, "max": 20} for m in muscles}


def sanitize_range_value(value, fallback):
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return fallback

    if numeric < 0:
        return fallback

    return numeric


def parse_requested_ranges(raw_ranges, muscles):
    defaults = build_default_ranges(muscles)
    if not isinstance(raw_ranges, dict):
        return defaults

    sanitized = {}
    for muscle in muscles:
        fallback = defaults[muscle]
        entry = raw_ranges.get(muscle, fallback)
        if not isinstance(entry, dict):
            sanitized[muscle] = fallback
            continue

        min_value = sanitize_range_value(entry.get("min"), fallback["min"])
        max_value = sanitize_range_value(entry.get("max"), fallback["max"])

        if max_value < min_value:
            max_value = min_value

        sanitized[muscle] = {"min": min_value, "max": max_value}

    return sanitized

@volume_splitter_bp.route('/volume_splitter')
def volume_splitter():
    requested_mode = (request.args.get('mode') or 'basic').lower()
    default_mode = 'advanced' if requested_mode == 'advanced' else 'basic'
    basic = BASIC_MUSCLE_GROUPS
    advanced = ADVANCED_MUSCLE_GROUPS
    return render_template(
        'volume_splitter.html',
        basic_muscle_groups=basic,
        advanced_muscle_groups=advanced,
        basic_recommended_ranges=build_default_ranges(basic),
        advanced_recommended_ranges=build_default_ranges(advanced),
        default_mode=default_mode
    )

@volume_splitter_bp.route('/api/calculate_volume', methods=['POST'])
def calculate_volume():
    data = request.get_json() or {}
    mode = (data.get('mode') or 'basic').lower()

    try:
        training_days = int(data.get('training_days', 3))
    except (TypeError, ValueError):
        training_days = 3
    training_days = max(training_days, 1)

    volumes = data.get('volumes', {}) or {}

    active_muscles = get_muscle_list_for_mode(mode)
    valid_muscles = set(active_muscles)
    requested_ranges = data.get('ranges') or {}
    ranges = parse_requested_ranges(requested_ranges, active_muscles)

    results = {}
    for muscle, weekly_sets in volumes.items():
        if muscle not in valid_muscles:
            continue

        try:
            weekly_sets_value = float(weekly_sets or 0)
        except (TypeError, ValueError):
            weekly_sets_value = 0.0

        sets_per_session = round(weekly_sets_value / training_days, 1)
        status = 'optimal'

        if weekly_sets_value < ranges[muscle]['min']:
            status = 'low'
        elif weekly_sets_value > ranges[muscle]['max']:
            status = 'high'

        if sets_per_session > 10:
            status = 'excessive'

        results[muscle] = {
            'weekly_sets': weekly_sets_value,
            'sets_per_session': sets_per_session,
            'status': status
        }

    suggestions = generate_volume_suggestions(training_days, results, mode=mode)

    return jsonify({'results': results, 'suggestions': suggestions, 'ranges': ranges})

@volume_splitter_bp.route('/api/volume_history')
def get_volume_history():
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT vp.id, vp.training_days, vp.created_at,
                   mv.muscle_group, mv.weekly_sets, mv.sets_per_session, mv.status
            FROM volume_plans vp
            JOIN muscle_volumes mv ON vp.id = mv.plan_id
            ORDER BY vp.created_at DESC
            LIMIT 100
        ''')
        history = cursor.fetchall()
        
        # Format the data for frontend
        formatted_history = {}
        for row in history:
            plan_id = row[0]
            if plan_id not in formatted_history:
                formatted_history[plan_id] = {
                    'training_days': row[1],
                    'created_at': row[2],
                    'muscles': {}
                }
            formatted_history[plan_id]['muscles'][row[3]] = {
                'weekly_sets': row[4],
                'sets_per_session': row[5],
                'status': row[6]
            }
        
        return jsonify(formatted_history)
        
    finally:
        conn.close()

@volume_splitter_bp.route('/api/save_volume_plan', methods=['POST'])
def save_volume_plan():
    data = request.get_json()
    plan_id = export_volume_plan(data)
    
    if plan_id:
        return jsonify({'success': True, 'plan_id': plan_id})
    return jsonify({'success': False, 'error': 'Failed to save plan'}), 500 

@volume_splitter_bp.route('/api/volume_plan/<int:plan_id>')
def get_volume_plan(plan_id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        # Get plan details
        cursor.execute('''
            SELECT vp.*, mv.muscle_group, mv.weekly_sets, mv.sets_per_session, mv.status
            FROM volume_plans vp
            JOIN muscle_volumes mv ON vp.id = mv.plan_id
            WHERE vp.id = ?
        ''', (plan_id,))
        
        rows = cursor.fetchall()
        if not rows:
            return jsonify({'error': 'Plan not found'}), 404
            
        plan = {
            'training_days': rows[0][1],
            'created_at': rows[0][2],
            'volumes': {}
        }
        
        for row in rows:
            plan['volumes'][row[3]] = {
                'weekly_sets': row[4],
                'sets_per_session': row[5],
                'status': row[6]
            }
            
        return jsonify(plan)
        
    finally:
        conn.close() 

@volume_splitter_bp.route('/api/volume_plan/<int:plan_id>', methods=['DELETE'])
def delete_volume_plan(plan_id):
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        # Check if plan exists
        cursor.execute('SELECT id FROM volume_plans WHERE id = ?', (plan_id,))
        if not cursor.fetchone():
            return jsonify({'success': False, 'error': 'Plan not found'}), 404
        
        # Delete the plan (muscle_volumes will cascade delete)
        cursor.execute('DELETE FROM volume_plans WHERE id = ?', (plan_id,))
        conn.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()

@volume_splitter_bp.route('/api/export_volume_excel', methods=['POST'])
def export_volume_excel():
    data = request.get_json() or {}
    try:
        training_days = int(data.get('training_days', 3))
    except (TypeError, ValueError):
        training_days = 3
    training_days = max(training_days, 1)
    volumes = data.get('volumes', {})
    
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Volume Plan"
    
    # Add headers
    headers = ['Muscle Group', 'Sets per Week', 'Sets per Session']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add data
    row = 2
    for muscle, weekly_sets in volumes.items():
        sets_per_session = round(weekly_sets / training_days, 1)
        ws.cell(row=row, column=1, value=muscle)
        ws.cell(row=row, column=2, value=weekly_sets)
        ws.cell(row=row, column=3, value=sets_per_session)
        row += 1
    
    # Style the worksheet
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    # Create the file
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'volume_plan_{timestamp}.xlsx'
    ) 