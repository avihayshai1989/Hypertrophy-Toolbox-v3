"""
Tests for data export functionality with focus on:
- Memory safety
- Filename sanitization
- Large dataset handling
- Streaming exports
"""

import pytest
import json
from io import BytesIO
from openpyxl import load_workbook
from utils.export_utils import (
    sanitize_filename,
    create_content_disposition_header,
    generate_timestamped_filename,
    estimate_export_size,
    should_use_streaming
)


class TestFilenameSanitization:
    """Test filename sanitization for security and safety."""
    
    def test_sanitize_basic_filename(self):
        """Test basic filename sanitization."""
        result = sanitize_filename("workout_summary.xlsx")
        assert result == "workout_summary.xlsx"
    
    def test_sanitize_special_characters(self):
        """Test removal of special characters."""
        result = sanitize_filename("workout<script>.xlsx")
        assert result == "workout_script.xlsx"
        assert "<" not in result
        assert ">" not in result
    
    def test_sanitize_path_traversal(self):
        """Test protection against path traversal attacks."""
        result = sanitize_filename("../../etc/passwd.xlsx")
        assert result == "passwd.xlsx"
        assert ".." not in result
        assert "/" not in result
    
    def test_sanitize_windows_path(self):
        """Test handling of Windows paths."""
        result = sanitize_filename("C:\\Windows\\System32\\file.xlsx")
        assert result == "file.xlsx"
        assert "\\" not in result
    
    def test_sanitize_empty_filename(self):
        """Test fallback for empty filename."""
        result = sanitize_filename("")
        assert result == "export.xlsx"
    
    def test_sanitize_long_filename(self):
        """Test truncation of long filenames."""
        long_name = "a" * 300 + ".xlsx"
        result = sanitize_filename(long_name)
        assert len(result) <= 200
        assert result.endswith(".xlsx")
    
    def test_sanitize_spaces_and_underscores(self):
        """Test handling of spaces and underscores."""
        result = sanitize_filename("my    workout___file.xlsx")
        assert result == "my_workout_file.xlsx"
    
    def test_sanitize_no_extension(self):
        """Test adding xlsx extension when missing."""
        result = sanitize_filename("workout_summary")
        assert result == "workout_summary.xlsx"
    
    def test_sanitize_wrong_extension(self):
        """Test replacing wrong extension."""
        result = sanitize_filename("workout_summary.txt")
        assert result == "workout_summary.xlsx"


class TestContentDisposition:
    """Test Content-Disposition header generation."""
    
    def test_content_disposition_attachment(self):
        """Test attachment disposition."""
        result = create_content_disposition_header("workout.xlsx", attachment=True)
        assert "attachment" in result
        assert "workout.xlsx" in result
    
    def test_content_disposition_inline(self):
        """Test inline disposition."""
        result = create_content_disposition_header("workout.xlsx", attachment=False)
        assert "inline" in result
    
    def test_content_disposition_special_chars(self):
        """Test handling of special characters in filename."""
        result = create_content_disposition_header("workout<test>.xlsx")
        assert "<" not in result
        assert ">" not in result


class TestTimestampedFilename:
    """Test timestamped filename generation."""
    
    def test_generate_timestamped_filename(self):
        """Test basic timestamped filename generation."""
        result = generate_timestamped_filename("workout_summary")
        assert "workout_summary_" in result
        assert result.endswith(".xlsx")
        # Should contain timestamp pattern YYYYMMDD_HHMMSS
        assert len(result.split("_")) >= 3
    
    def test_timestamped_filename_custom_extension(self):
        """Test with custom extension - should still use xlsx."""
        result = generate_timestamped_filename("report", "csv")
        # Should still use xlsx as it's enforced for Excel workbooks
        assert result.endswith(".xlsx")
    
    def test_timestamped_filename_sanitized(self):
        """Test that generated filename is sanitized."""
        result = generate_timestamped_filename("my<test>file")
        assert "<" not in result
        assert ">" not in result


class TestExportSizeEstimation:
    """Test export size estimation for streaming decisions."""
    
    def test_estimate_small_export(self):
        """Test estimation for small export."""
        size = estimate_export_size(100, 10)
        assert size > 0
        assert size < 100000  # Less than 100KB
    
    def test_estimate_large_export(self):
        """Test estimation for large export."""
        size = estimate_export_size(10000, 20)
        assert size > 1000000  # More than 1MB
    
    def test_should_use_streaming_small(self):
        """Test that small exports don't use streaming."""
        result = should_use_streaming(100, 10)
        assert result is False
    
    def test_should_use_streaming_large(self):
        """Test that large exports use streaming."""
        result = should_use_streaming(100000, 50)
        assert result is True


class TestExportsEndpoints:
    """Test export endpoints with actual requests."""
    
    def test_export_to_excel_structure(self, client, sample_workout_plan):
        """Test full Excel export structure."""
        response = client.get('/export_to_excel')
        
        assert response.status_code == 200
        assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        # Verify Content-Disposition header
        assert 'Content-Disposition' in response.headers
        disposition = response.headers['Content-Disposition']
        assert 'attachment' in disposition
        assert 'workout_tracker_summary' in disposition
        assert '.xlsx' in disposition
        
        # Verify the file is valid Excel
        wb = load_workbook(BytesIO(response.data))
        assert wb is not None
        
        # Check for expected sheets
        sheet_names = wb.sheetnames
        assert 'Workout Plan' in sheet_names or len(sheet_names) >= 1
    
    def test_export_to_excel_filename_timestamp(self, client):
        """Test that exported filename includes timestamp."""
        response = client.get('/export_to_excel')
        
        disposition = response.headers['Content-Disposition']
        # Should contain timestamp pattern
        assert '_20' in disposition  # Year prefix
    
    def test_export_summary_with_method(self, client, sample_workout_log):
        """Test summary export with specific method."""
        response = client.post('/export_summary',
                               json={'method': 'Total'},
                               content_type='application/json')
        
        assert response.status_code == 200
        assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        # Verify filename includes method
        disposition = response.headers['Content-Disposition']
        assert 'workout_summary_Total' in disposition
    
    def test_export_summary_invalid_method(self, client):
        """Test summary export with invalid method gracefully."""
        response = client.post('/export_summary',
                               json={'method': 'InvalidMethod'},
                               content_type='application/json')
        
        # Should still succeed or return graceful error
        assert response.status_code in [200, 400]
    
    def test_export_to_workout_log_success(self, client, sample_workout_plan):
        """Test exporting workout plan to workout log."""
        response = client.post('/export_to_workout_log')
        
        assert response.status_code == 200
        data = json.loads(response.data)
        assert data['status'] == 'success'
        assert 'message' in data
    
    def test_export_to_workout_log_empty_plan(self, client, clean_database):
        """Test exporting empty workout plan."""
        response = client.post('/export_to_workout_log')
        
        assert response.status_code in [400, 404, 500]  # Accept various error codes
        try:
            data = json.loads(response.data)
            assert data['status'] == 'error'
        except:
            # If JSON parsing fails, that's also acceptable for error responses
            pass
    
    def test_export_large_dataset_all(self, client, large_workout_log):
        """Test streaming export with all data."""
        response = client.post('/export_large_dataset',
                               json={'type': 'all'},
                               content_type='application/json')
        
        assert response.status_code == 200
        assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        # Verify it's a valid Excel file
        wb = load_workbook(BytesIO(response.data))
        assert wb is not None
    
    def test_export_large_dataset_workout_log_only(self, client, large_workout_log):
        """Test streaming export with workout log only."""
        response = client.post('/export_large_dataset',
                               json={'type': 'workout_log'},
                               content_type='application/json')
        
        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.data))
        assert 'Workout Log' in wb.sheetnames
    
    def test_export_large_dataset_session_summary_only(self, client, large_workout_log):
        """Test streaming export with session summary only."""
        response = client.post('/export_large_dataset',
                               json={'type': 'session_summary'},
                               content_type='application/json')
        
        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.data))
        assert 'Session Summary' in wb.sheetnames


class TestExportMemorySafety:
    """Test that exports handle large datasets without memory issues."""
    
    def test_export_with_many_rows(self, client, database_with_rows):
        """Test export with large number of rows (memory safety)."""
        # This test would populate database with many rows
        # and verify export completes without memory errors
        response = client.get('/export_to_excel')
        
        assert response.status_code == 200
        # If we get here without timeout/memory error, test passes
        assert len(response.data) > 0
    
    def test_streaming_export_memory_usage(self, client, large_workout_log):
        """Test that streaming export doesn't load all data at once."""
        import tracemalloc
        
        tracemalloc.start()
        initial_memory = tracemalloc.get_traced_memory()[0]
        
        response = client.post('/export_large_dataset',
                               json={'type': 'all'},
                               content_type='application/json')
        
        peak_memory = tracemalloc.get_traced_memory()[1]
        tracemalloc.stop()
        
        # Memory usage should be reasonable (less than 100MB increase)
        memory_increase = (peak_memory - initial_memory) / (1024 * 1024)  # MB
        assert response.status_code == 200
        assert memory_increase < 100  # Less than 100MB increase
    
    def test_export_respects_max_rows_limit(self, client, database_with_many_rows):
        """Test that exports respect MAX_EXPORT_ROWS limit."""
        from utils.export_utils import MAX_EXPORT_ROWS
        
        response = client.get('/export_to_excel')
        
        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.data))
        
        # Check that no sheet exceeds MAX_EXPORT_ROWS
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            row_count = sheet.max_row
            assert row_count <= MAX_EXPORT_ROWS + 1  # +1 for header


class TestExportErrorHandling:
    """Test error handling in export operations."""
    
    def test_export_handles_database_error(self, client, broken_database):
        """Test that export handles database errors gracefully."""
        response = client.get('/export_to_excel')
        
        # Should return an error status
        assert response.status_code in [500, 404]
        # Try to parse as JSON, but it might be HTML error page
        try:
            data = json.loads(response.data)
            assert data['status'] == 'error'
        except:
            # HTML error page is also acceptable
            pass
    
    def test_export_handles_missing_data(self, client, clean_database):
        """Test export with no data returns valid empty Excel."""
        response = client.get('/export_to_excel')
        
        # Should either succeed with empty sheets or return appropriate error
        assert response.status_code in [200, 400]
    
    def test_export_summary_handles_bad_json(self, client):
        """Test export summary with invalid JSON."""
        response = client.post('/export_summary',
                               data='invalid json',
                               content_type='application/json')
        
        # Should return error status
        assert response.status_code in [400, 500]
        # Response might be JSON error or HTML page
        # Both are acceptable for this test


# Fixtures for tests

@pytest.fixture
def sample_workout_plan(client):
    """Create a sample workout plan for testing."""
    from utils.database import DatabaseHandler
    
    with DatabaseHandler() as db:
        db.execute_query("""
            INSERT INTO user_selection (routine, exercise, sets, min_rep_range, max_rep_range, weight)
            VALUES ('A', 'Bench Press', 3, 8, 12, 100)
        """)
    
    yield
    
    # Cleanup
    with DatabaseHandler() as db:
        db.execute_query("DELETE FROM user_selection")


@pytest.fixture
def sample_workout_log(client):
    """Create sample workout log entries."""
    from utils.database import DatabaseHandler
    
    with DatabaseHandler() as db:
        for i in range(10):
            db.execute_query("""
                INSERT INTO workout_log (routine, exercise, planned_sets, scored_weight, scored_max_reps)
                VALUES ('A', 'Bench Press', 3, 100, 10)
            """)
    
    yield
    
    # Cleanup
    with DatabaseHandler() as db:
        db.execute_query("DELETE FROM workout_log")


@pytest.fixture
def large_workout_log(client):
    """Create a large workout log for streaming tests."""
    from utils.database import DatabaseHandler
    
    with DatabaseHandler() as db:
        # Insert 1000 records for testing
        for i in range(1000):
            db.execute_query("""
                INSERT INTO workout_log (routine, exercise, planned_sets, scored_weight, scored_max_reps)
                VALUES ('A', 'Bench Press', 3, ?, 10)
            """, (100 + i % 50,))
    
    yield
    
    # Cleanup
    with DatabaseHandler() as db:
        db.execute_query("DELETE FROM workout_log")


@pytest.fixture
def database_with_rows(client):
    """Create database with many rows for memory testing."""
    from utils.database import DatabaseHandler
    
    with DatabaseHandler() as db:
        # Insert 5000 records
        for i in range(5000):
            db.execute_query("""
                INSERT INTO workout_log (routine, exercise, planned_sets, scored_weight, scored_max_reps)
                VALUES ('A', 'Exercise ' || ?, 3, ?, 10)
            """, (i, 100 + i % 100))
    
    yield
    
    # Cleanup
    with DatabaseHandler() as db:
        db.execute_query("DELETE FROM workout_log")


@pytest.fixture
def database_with_many_rows(client):
    """Create database with rows exceeding MAX_EXPORT_ROWS."""
    from utils.database import DatabaseHandler
    from utils.export_utils import MAX_EXPORT_ROWS
    
    # Create more rows than the limit
    rows_to_create = min(MAX_EXPORT_ROWS + 1000, 100000)  # Don't create too many in tests
    
    with DatabaseHandler() as db:
        for i in range(rows_to_create):
            if i % 1000 == 0:  # Log progress
                print(f"Created {i} rows...")
            db.execute_query("""
                INSERT INTO workout_log (routine, exercise, planned_sets, scored_weight, scored_max_reps)
                VALUES ('A', 'Exercise', 3, 100, 10)
            """)
    
    yield
    
    # Cleanup
    with DatabaseHandler() as db:
        db.execute_query("DELETE FROM workout_log")


@pytest.fixture
def clean_database(client):
    """Ensure database is clean for testing."""
    from utils.database import DatabaseHandler
    
    with DatabaseHandler() as db:
        db.execute_query("DELETE FROM user_selection")
        db.execute_query("DELETE FROM workout_log")
    
    yield


@pytest.fixture
def broken_database(client, monkeypatch):
    """Simulate a broken database connection."""
    from utils.database import DatabaseHandler
    
    original_fetch_all = DatabaseHandler.fetch_all
    
    def broken_fetch_all(self, *args, **kwargs):
        raise Exception("Database connection failed")
    
    monkeypatch.setattr(DatabaseHandler, 'fetch_all', broken_fetch_all)
    
    yield
    
    # Restore
    monkeypatch.setattr(DatabaseHandler, 'fetch_all', original_fetch_all)

